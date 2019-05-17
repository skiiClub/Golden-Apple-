import openpyxl, selenium
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time 
#reads input excel file 
wb = openpyxl.load_workbook('C:/Users/Chomiak/Documents/Golden Apple/tabula-schoolsByNetwork.xlsx', data_only=True)

sheet = wb.get_sheet_by_name('tabula-schoolsByNetwork')


#Gather schools in data sheet 

schoolList = []
for i in range(1,sheet.max_row):
    for j in range (1,sheet.max_column):
        school = sheet.cell(row=i, column=j).value
        if str(school) != 'None':
            schoolList.append(school)




#open website

browser = webdriver.Firefox()

browser.get('http://cps.edu/ScriptLibrary/Map-SchoolLocator/index.html#')

#create excel sheet to export data to 
new_wb = openpyxl.Workbook()
sheet = new_wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'School'
sheet['B1'] = 'Performance Rating'

#for columns 
sheet_incrementer = 2

#iterate through schools 
for school in schoolList:
    searchElem = browser.find_element_by_name('search')
    searchElem.send_keys(school)
    buttonElem = browser.find_element_by_class_name('btnSearch')
    buttonElem.click()
    #fixes random error with div 
    time.sleep(1)

    #Scraping the website for the performance rating. 
    html = browser.page_source
    soup = BeautifulSoup(html, "lxml")
    #sorting with collapse in, returns two divs. 
    elems = soup.find_all('div', attrs={'class': 'collapse in'})
    toParse = str(elems[0]) + "parse here: " +str(elems[1]) 
    result = toParse.find('level')
    end = result + 7
    #if the search can't find the school this will break
    try:
        test = int(toParse[result+6])

    except ValueError:
        print(school+ " needs to be entered manually\n")
        test = 1 
    #writing to the excel document 
    if test >=2:
        a = 'A'+str(sheet_incrementer)
        b = 'B'+str(sheet_incrementer)
        sheet_incrementer = sheet_incrementer + 1
    #handling if toParse is + or ' '
        if toParse[end] == '+':
            end = end + 1
        sheet[a]= school
        sheet[b]= toParse[result:end]
    #clearing the entre and continuing the iteration 
    eraseElem = browser.find_element_by_class_name('btnClearSearch')
    eraseElem.click()


#saving data and closing
browser.quit()
new_wb.save('results.xlsx')



